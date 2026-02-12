import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
from datetime import datetime

try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    print("Le module openpyxl est requis. Veuillez l'installer avec: pip install openpyxl")
    sys.exit(1)

# Family code to extensions mapping - easy to modify
FAMILY_CODE_EXTENSIONS = {
    "FAM0201": ["igs", "stp"],      # Laser tube inf 5850 mm
    "FAM0202": ["igs", "stp"],      # Laser tube sup 5850 mm
    "FAM0203": ["dxf"],             # Tole inf 3000 mm
    "FAM0204": ["dxf"],             # Tole sup 3000 mm
    "FAM0206": ["step"],            # Tole pliée gde long
    "FAM0207": ["step"],            # Tole pliée spé
    "FAM0208": ["step"],            # Tole pliée
}

# Base extensions required for all files
BASE_EXTENSIONS = ["slddrw", "pdf"]


def find_nomenclature_excel(folder_path):
    """
    Automatically find Excel file starting with 'Nomenclature_' in the folder.
    """
    try:
        files = os.listdir(folder_path)
        excel_files = [f for f in files 
                      if f.lower().startswith('nomenclature_') and f.lower().endswith(('.xlsx', '.xls'))]
        
        if excel_files:
            # Return the first matching Excel file
            return os.path.join(folder_path, excel_files[0])
        return None
    except Exception as e:
        print(f"Error searching for Excel file: {e}")
        return None


def load_extensions_from_excel(excel_path, folder_path=None):
    """
    Load file extensions requirements from an Excel file.
    Expected format: Column A = filename, Column C = family code, Column D = file path.
    Only includes files that are in the specified folder.
    """
    try:
        workbook = load_workbook(excel_path)
        if workbook is None:
            messagebox.showerror("Erreur", f"Impossible de charger le fichier Excel: {excel_path}")
            return None
            
        sheet = workbook.active
        if sheet is None:
            messagebox.showerror("Erreur", f"Aucune feuille trouvée dans le fichier Excel: {excel_path}")
            return None
        
        # Dictionary to store required extensions for each file
        file_extensions_map = {}
        
        # Skip header row
        first_row = True
        for row in sheet.iter_rows(values_only=True):
            if first_row:
                first_row = False
                continue
            
            if row and len(row) >= 4 and row[0] and row[2]:
                filename = str(row[0]).strip()
                family_code = str(row[2]).strip().upper()
                file_path = str(row[3]).strip() if len(row) >= 4 and row[3] else ""
                
                # Skip files with N/A as family code
                if family_code == "N/A":
                    continue
                
                # Only check files that are in the specified folder
                if folder_path and file_path:
                    # Normalize paths for comparison
                    normalized_file_path = os.path.normpath(file_path)
                    normalized_folder_path = os.path.normpath(folder_path)
                    
                    # Check if the file path matches the folder path
                    if not normalized_file_path.lower().startswith(normalized_folder_path.lower()):
                        continue
                
                # Determine required extensions based on family code
                required_extensions = BASE_EXTENSIONS.copy()  # All files need these
                
                # Add family-specific extensions
                for code, extensions in FAMILY_CODE_EXTENSIONS.items():
                    if code in family_code:
                        required_extensions.extend(extensions)
                        break
                
                file_extensions_map[filename] = required_extensions
        
        return file_extensions_map
        
    except Exception as e:
        messagebox.showerror("Erreur", f"Erreur lors de la lecture du fichier Excel: {e}")
        return None


def check_folder(folder_path, text_widget, status_label, excel_path=None):
    report = {
        "fichiers_manquants": [],
        "fichiers_obsolètes": [],
        "fichiers_non_verifies": [],
    }

    # Custom titles for each report section
    report_titles = {
        "fichiers_manquants": "Fichiers manquants selon les règles de famille",
        "fichiers_obsolètes": "Fichiers obsolètes (le fichier source est plus récent que le fichier exporté)",
        "fichiers_non_verifies": "Fichiers non vérifiés (ne correspondent à aucune règle)",
    }

    # Load extensions from Excel if provided, or auto-detect
    file_extensions_map = {}
    used_excel_path = None
    
    # Try to use provided Excel path first
    if excel_path and os.path.exists(excel_path):
        file_extensions_map = load_extensions_from_excel(excel_path, folder_path)
        used_excel_path = excel_path
        if file_extensions_map is None:
            return  # Error already shown
    else:
        # Auto-detect Excel file if not provided
        auto_excel_path = find_nomenclature_excel(folder_path)
        if auto_excel_path and os.path.exists(auto_excel_path):
            file_extensions_map = load_extensions_from_excel(auto_excel_path, folder_path)
            used_excel_path = auto_excel_path
            if file_extensions_map is None:
                return  # Error already shown
        else:
            # Default extensions if no Excel file found
            file_extensions_map = {
                "Piece_1": ["slddrw", "pdf", "igs", "step"],  # Example for FAM0201
                "Piece_2": ["slddrw", "pdf", "dxf"],       # Example for FAM0203
                "ENS-BOR-HUL-120-120-A": ["slddrw", "pdf"],  # Default rules for these files
                "ENS-EMB-BOR-HUL-120-A": ["slddrw", "pdf"],
                "ENS-TET-HUL-120-A": ["slddrw", "pdf"],
            }

    # Get all files in the folder (case-insensitive), skipping folders and Excel temporary files
    files_original = [f for f in os.listdir(folder_path) 
                     if os.path.isfile(os.path.join(folder_path, f)) and not f.startswith('~$')]
    files = [f.lower() for f in files_original]

    # Track which files were checked
    checked_files = set()

    # Check files according to file extensions map
    for filename, required_exts in file_extensions_map.items():
        base_name = filename
        
        # Check if required extensions exist
        missing_exts = []
        for req_ext in required_exts:
            required_file = f"{base_name.lower()}.{req_ext}"
            if required_file not in files and required_file.upper() not in [f.upper() for f in files_original]:
                missing_exts.append(req_ext)
        
        if missing_exts:
            report["fichiers_manquants"].append(
                f"{filename} (manque : {', '.join(missing_exts)})"
            )
        else:
            # Mark this file as checked
            for f in files_original:
                if f.lower().startswith(base_name.lower() + "."):
                    checked_files.add(f.lower())
            
            # Check for outdated files (if source file is newer than exported files)
            # Look for SLDPRT or SLDASM file with this base name
            source_files = [f for f in files_original 
                          if f.lower().startswith(base_name.lower() + ".") 
                          and f.lower().endswith(('.sldprt', '.sldasm'))]
            
            for source_file in source_files:
                source_path = os.path.join(folder_path, source_file)
                source_mtime = os.path.getmtime(source_path)
                
                for req_ext in required_exts:
                    required_file = f"{base_name.lower()}.{req_ext}"
                    if required_file in files or required_file.upper() in [f.upper() for f in files_original]:
                        # Find the actual file (case insensitive)
                        actual_file = None
                        for f in files_original:
                            if f.lower() == required_file or f.upper() == required_file.upper():
                                actual_file = f
                                break
                        
                        if actual_file:
                            exported_path = os.path.join(folder_path, actual_file)
                            exported_mtime = os.path.getmtime(exported_path)
                            
                            if source_mtime > exported_mtime:
                                report["fichiers_obsolètes"].append(
                                    f"{source_file} (plus récent que {actual_file})"
                                )

    # Check for unchecked files
    all_files_lower = [f.lower() for f in files_original]
    for file in files_original:
        if file.lower() not in checked_files:
            # Skip Excel files, common non-design files, and SLDPRT/SLDASM/SLDDRW files (these are the source files being verified)
            file_upper = file.upper()
            if (not file_upper.endswith('.SLDPRT') and 
                not file_upper.endswith('.SLDASM') and 
                not file_upper.endswith('.SLDDRW') and
                not any(file_upper.endswith(ext) for ext in ['.XLSX', '.XLS', '.CSV', '.BAT', '.ICO', '.REG'])):
                report["fichiers_non_verifies"].append(file)

    # Update UI
    text_widget.delete(1.0, tk.END)
    all_ok = True

    for issue, missing_files in report.items():
        if issue != "fichiers_non_verifies" and missing_files:
            all_ok = False
            break

    # Set status label
    if all_ok:
        status_label.config(
            text="✅ Tous les fichiers vérifiés sont présents !", foreground="green"
        )
    else:
        status_label.config(
            text="❌ Des fichiers vérifiés sont manquants.", foreground="red"
        )

    # Insert report with custom titles
    text_widget.insert(tk.END, "=== Rapport de vérification ===\n\n", "title")

    # Show which Excel file was used
    if used_excel_path:
        text_widget.insert(tk.END, f"Fichier de règles utilisé: {os.path.basename(used_excel_path)}\n\n")
    else:
        text_widget.insert(tk.END, "Utilisation des règles par défaut\n\n")

    for issue, missing_files in report.items():
        # Use custom title from report_titles
        text_widget.insert(tk.END, f"{report_titles[issue]} :\n", "section")

        if missing_files:
            for file in missing_files:
                text_widget.insert(
                    tk.END,
                    f"  - {file}\n",
                    (
                        "missing"
                        if issue not in ["fichiers_non_verifies", "fichiers_obsolètes"]
                        else (
                            "unverified"
                            if issue == "fichiers_non_verifies"
                            else "warning"
                        )
                    ),
                )
        else:
            text_widget.insert(tk.END, "  ✅ 0 fichier manquant\n", "ok")

        text_widget.insert(tk.END, "\n")

    # Configure text tags for colors
    text_widget.tag_config("title", font=("Arial", 12, "bold"))
    text_widget.tag_config("section", font=("Arial", 10, "bold"))
    text_widget.tag_config("ok", foreground="green")
    text_widget.tag_config("missing", foreground="red")
    text_widget.tag_config("unverified", foreground="orange")
    text_widget.tag_config("warning", foreground="purple")  # Purple for outdated files


def select_folder(text_widget, status_label, excel_path_var):
    folder_path = filedialog.askdirectory(title="Sélectionner un dossier")
    if folder_path:
        check_folder(folder_path, text_widget, status_label, excel_path_var.get())
    else:
        messagebox.showwarning("Avertissement", "Aucun dossier sélectionné.")


def select_excel(excel_path_var):
    excel_path = filedialog.askopenfilename(
        title="Sélectionner un fichier Excel",
        filetypes=[("Fichiers Excel", "*.xlsx;*.xls"), ("Tous les fichiers", "*.*")]
    )
    if excel_path:
        excel_path_var.set(excel_path)


def main():
    root = tk.Tk()

    # Get folder path from command line argument
    if len(sys.argv) > 1:
        folder_path = sys.argv[1]
        folder_name = os.path.basename(folder_path)
        root.title(f"Vérificateur de Dossier - {folder_name}")
    else:
        root.title("Vérificateur de Dossier")
        folder_path = None

    # Excel path variable
    excel_path_var = tk.StringVar()

    # Frame for controls
    control_frame = tk.Frame(root)
    control_frame.pack(pady=5, fill=tk.X)

    # Excel file selection
    excel_frame = tk.Frame(control_frame)
    excel_frame.pack(side=tk.LEFT, padx=5)

    excel_label = tk.Label(excel_frame, text="Fichier Excel:")
    excel_label.pack(side=tk.LEFT)

    excel_entry = tk.Entry(excel_frame, textvariable=excel_path_var, width=50)
    excel_entry.pack(side=tk.LEFT, padx=5)

    excel_button = tk.Button(excel_frame, text="Parcourir", command=lambda: select_excel(excel_path_var))
    excel_button.pack(side=tk.LEFT)

    # Folder selection button
    folder_button = tk.Button(control_frame, text="Sélectionner Dossier", 
                             command=lambda: select_folder(text_widget, status_label, excel_path_var))
    folder_button.pack(side=tk.RIGHT, padx=5)

    # Label to display the folder path at the top
    folder_label = tk.Label(
        root,
        text=f"Dossier: {folder_path if folder_path else 'Aucun dossier sélectionné'}",
        font=("Arial", 10, "bold"),
    )
    folder_label.pack(pady=5)

    # Status label
    status_label = tk.Label(root, text="Vérification en cours...", font=("Arial", 10))
    status_label.pack(pady=5)

    # Text widget to display the report
    text_widget = scrolledtext.ScrolledText(
        root, width=100, height=50, wrap=tk.WORD, font=("Arial", 10)
    )
    text_widget.pack(padx=10, pady=10)

    # Run the check if a folder path is provided
    if folder_path:
        check_folder(folder_path, text_widget, status_label, excel_path_var.get())
    else:
        text_widget.insert(tk.END, "Aucun chemin de dossier fourni en argument.\n")
        status_label.config(text="❌ Aucun chemin de dossier fourni.", foreground="red")

    root.mainloop()


if __name__ == "__main__":
    main()